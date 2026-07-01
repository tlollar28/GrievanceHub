from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def parse_dt(dt_string: str):
    return datetime.strptime(dt_string, "%Y-%m-%d %H:%M")


def overlaps(start1, end1, start2, end2):
    return start1 < end2 and start2 < end1


def generate_findings(
    runtime_sessions,
    clerks_assignments,
    pse_assignments,
    mail_handler_assignments,
    allowed_clerks=0,
):
    findings = []

    all_assignments = (
        clerks_assignments
        + pse_assignments
        + mail_handler_assignments
    )

    for session in runtime_sessions:
        session_start = parse_dt(session["start_time"])
        session_end = parse_dt(session["end_time"])

        clerks_present = []
        pses_present = []
        mail_handlers_present = []

        for assignment in all_assignments:
            assignment_start = parse_dt(assignment["start_time"])
            assignment_end = parse_dt(assignment["end_time"])

            if not overlaps(
                session_start,
                session_end,
                assignment_start,
                assignment_end,
            ):
                continue

            if assignment["craft"] == "clerk":
                if assignment["employee_type"] == "career_clerk":
                    clerks_present.append(assignment["employee_name"])
                else:
                    pses_present.append(assignment["employee_name"])

            elif assignment["craft"] == "mail_handler":
                mail_handlers_present.append(assignment["employee_name"])

        clerk_count = len(clerks_present)
        pse_count = len(pses_present)
        total_clerk_side_count = clerk_count + pse_count
        mail_handler_count = len(mail_handlers_present)

        required_mail_handlers = 0

        if total_clerk_side_count > allowed_clerks:
            status = "Potential Cross-Craft Grievance"
            reason = (
                f"{total_clerk_side_count} clerk-side employees exceeds "
                f"allowed {allowed_clerks}"
            )
        else:
            status = "Compliant"
            reason = "Within staffing limits"

        findings.append(
            {
                "date": session["date"],
                "machine": session["machine"],
                "machine_start_time": session["start_time"],
                "machine_end_time": session["end_time"],
                "duration_minutes": session["duration_minutes"],
                "clerk_names": clerks_present,
                "pse_names": pses_present,
                "mail_handler_names": mail_handlers_present,
                "clerk_count": clerk_count,
                "pse_count": pse_count,
                "total_clerk_side_count": total_clerk_side_count,
                "mail_handler_count": mail_handler_count,
                "allowed_clerks": allowed_clerks,
                "required_mail_handlers": required_mail_handlers,
                "status": status,
                "reason": reason,
            }
        )

    return findings


def create_cross_craft_report(report_path: str, findings: list):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Cross Craft Report"

    headers = [
        "Date",
        "Machine",
        "Code",
        "Start Time",
        "End Time",
        "Allowed Clerks",
        "Actual Clerks",
        "Overage",
        "Status",
    ]

    sheet.append(headers)

    for finding in findings:
        sheet.append(
            [
                finding["date"],
                finding["machine"],
                finding.get("code", ""),
                finding["machine_start_time"],
                finding["machine_end_time"],
                finding["allowed_clerks"],
                finding["total_clerk_side_count"],
                max(
                    finding["total_clerk_side_count"]
                    - finding["allowed_clerks"],
                    0,
                ),
                finding["status"],
            ]
        )

    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid",
    )
    bold_font = Font(bold=True)

    for cell in sheet[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        sheet.column_dimensions[column_letter].width = max_length + 3

    workbook.save(report_path)

    return report_path